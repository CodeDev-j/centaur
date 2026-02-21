"""
Excel Builder: Traceable workbook export with HYPERLINK cells.

Generates a formatted .xlsx workbook from metric_facts where every value
cell links back to the source page in Centaur. Pivot-ready format with
frozen headers, auto-filter, and accounting number formats.

Sheets:
- Summary: Document metadata + audit findings count
- Metrics: Full fact table with =HYPERLINK() formulas
- Series Index: Unique series with metadata
"""

import io
import logging
import os
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import text as sql_text

from src.config import SystemConfig
from src.storage.analytics_driver import AnalyticsDriver

logger = logging.getLogger(__name__)

# Configurable base URL for hyperlinks
BASE_URL = os.getenv("CENTAUR_BASE_URL", "http://localhost:3000")

# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINK_FONT = Font(name="Calibri", size=10, color="4F46E5", underline="single")
NUMBER_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.0%'


def build_workbook(
    doc_hash: str,
    analytics: AnalyticsDriver,
    doc_metadata: Optional[Dict[str, Any]] = None,
    audit_summary: Optional[Dict[str, int]] = None,
) -> io.BytesIO:
    """
    Generates a .xlsx workbook from metric_facts for a single document.

    Returns a BytesIO buffer ready for StreamingResponse.
    """
    wb = Workbook()

    # Fetch all metric facts for this document
    session = analytics.SessionLocal()
    try:
        facts = session.execute(
            sql_text("""
                SELECT page_number::int, series_label, series_nature,
                       label, numeric_value, currency, magnitude, measure,
                       resolved_value, period_date, accounting_basis,
                       data_provenance, periodicity, archetype,
                       confidence_score, original_text
                FROM metric_facts
                WHERE doc_hash = :hash
                ORDER BY page_number, series_label, period_date NULLS LAST, label
            """),
            {"hash": doc_hash},
        )
        rows = facts.fetchall()
        columns = facts.keys()

        # Fetch unique series
        series_result = session.execute(
            sql_text("""
                SELECT series_label, series_nature, periodicity,
                       accounting_basis, archetype, currency, magnitude,
                       COUNT(*) as data_points,
                       MIN(page_number)::int as first_page
                FROM metric_facts
                WHERE doc_hash = :hash
                GROUP BY series_label, series_nature, periodicity,
                         accounting_basis, archetype, currency, magnitude
                ORDER BY MIN(page_number), series_label
            """),
            {"hash": doc_hash},
        )
        series_rows = series_result.fetchall()
    finally:
        session.close()

    # ── Sheet 1: Summary ─────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    filename = doc_metadata.get("filename", "Unknown") if doc_metadata else "Unknown"
    summary_data = [
        ("Document", filename),
        ("Doc Hash", doc_hash[:16] + "..."),
        ("Total Metrics", len(rows)),
        ("Unique Series", len(series_rows)),
    ]

    if audit_summary:
        summary_data.extend([
            ("", ""),
            ("Audit Findings", ""),
            ("  Errors", audit_summary.get("error", 0)),
            ("  Warnings", audit_summary.get("warning", 0)),
            ("  Info", audit_summary.get("info", 0)),
        ])

    for i, (key, val) in enumerate(summary_data, 1):
        ws_summary.cell(row=i, column=1, value=key).font = Font(bold=True, size=10)
        ws_summary.cell(row=i, column=2, value=val)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40

    # ── Sheet 2: Metrics ─────────────────────────────────────────────
    ws_metrics = wb.create_sheet("Metrics")

    headers = [
        "Page", "Series", "Nature", "Period", "Value",
        "Currency", "Magnitude", "Resolved Value", "Measure",
        "Accounting Basis", "Periodicity", "Archetype",
        "Confidence", "Original Text", "Source",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws_metrics.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, row in enumerate(rows, 2):
        page = row[0]       # page_number
        series = row[1]     # series_label
        nature = row[2]     # series_nature
        label = row[3]      # label (period)
        num_val = row[4]    # numeric_value
        currency = row[5]   # currency
        magnitude = row[6]  # magnitude
        measure = row[7]    # measure
        resolved = row[8]   # resolved_value
        period_date = row[9]
        acct_basis = row[10]
        provenance = row[11]
        periodicity = row[12]
        archetype = row[13]
        confidence = row[14]
        original = row[15]

        ws_metrics.cell(row=row_idx, column=1, value=page)
        ws_metrics.cell(row=row_idx, column=2, value=series)
        ws_metrics.cell(row=row_idx, column=3, value=nature)
        ws_metrics.cell(row=row_idx, column=4, value=label)

        # Numeric value
        val_cell = ws_metrics.cell(row=row_idx, column=5, value=num_val)
        if num_val is not None:
            val_cell.number_format = NUMBER_FORMAT

        ws_metrics.cell(row=row_idx, column=6, value=currency if currency != "None" else "")
        ws_metrics.cell(row=row_idx, column=7, value=magnitude if magnitude != "None" else "")

        # Resolved value (key column — accounting format)
        resolved_cell = ws_metrics.cell(row=row_idx, column=8, value=resolved)
        if resolved is not None:
            resolved_cell.number_format = NUMBER_FORMAT

        ws_metrics.cell(row=row_idx, column=9, value=measure or "")
        ws_metrics.cell(row=row_idx, column=10, value=acct_basis or "")
        ws_metrics.cell(row=row_idx, column=11, value=periodicity or "")
        ws_metrics.cell(row=row_idx, column=12, value=archetype or "")

        conf_cell = ws_metrics.cell(row=row_idx, column=13, value=confidence)
        if confidence is not None:
            conf_cell.number_format = '0.00'

        ws_metrics.cell(row=row_idx, column=14, value=original or "")

        # Source hyperlink — links back to Centaur at the specific page
        link_url = f"{BASE_URL}?doc={doc_hash}&page={page}"
        source_cell = ws_metrics.cell(row=row_idx, column=15)
        source_cell.value = f"Page {page}"
        source_cell.hyperlink = link_url
        source_cell.font = LINK_FONT

    # Freeze top row + auto-filter
    ws_metrics.freeze_panes = "A2"
    ws_metrics.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Auto-width columns
    col_widths = [8, 25, 10, 15, 14, 10, 12, 18, 10, 16, 14, 14, 10, 30, 12]
    for i, w in enumerate(col_widths, 1):
        ws_metrics.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Series Index ────────────────────────────────────────
    ws_series = wb.create_sheet("Series Index")

    series_headers = [
        "Series Label", "Nature", "Periodicity", "Accounting Basis",
        "Archetype", "Currency", "Magnitude", "Data Points", "First Page",
    ]

    for col_idx, header in enumerate(series_headers, 1):
        cell = ws_series.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, row in enumerate(series_rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws_series.cell(row=row_idx, column=col_idx)
            if val is None or str(val) == "None":
                cell.value = ""
            else:
                cell.value = val

    ws_series.freeze_panes = "A2"
    ws_series.auto_filter.ref = (
        f"A1:{get_column_letter(len(series_headers))}{len(series_rows) + 1}"
    )

    series_widths = [30, 10, 14, 16, 14, 10, 12, 12, 12]
    for i, w in enumerate(series_widths, 1):
        ws_series.column_dimensions[get_column_letter(i)].width = w

    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        f"Excel workbook built: {len(rows)} metrics, "
        f"{len(series_rows)} series for {doc_hash[:8]}"
    )
    return buffer
