import logging
import hashlib
from pathlib import Path
from typing import List, Optional
from uuid import uuid4

# Third-party
from markitdown import MarkItDown
from openpyxl import load_workbook
from langsmith import traceable

# Internal
from src.schemas.documents import IngestedChunk
from src.storage.blob_driver import BlobDriver
from src.ingestion.phantom import PhantomRenderer # We will build this next

logger = logging.getLogger(__name__)

class NativeParser:
    """
    Helix B: The Native Stream.
    Processes Excel/Word/PPT using Microsoft MarkItDown.
    Includes 'Defensive Fallback' to OpenPyXL for messy spreadsheets.
    """
    
    def __init__(self):
        self.md_converter = MarkItDown()
        self.phantom = PhantomRenderer()

    @traceable(name="Parse Native File", run_type="parser")
    async def parse(self, file_path: Path) -> List[IngestedChunk]:
        """
        Main entry point. Routes to specific logic based on extension.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        logger.info(f"📊 Starting Native ingestion for: {file_path.name}")
        
        # 1. Generate Stable Hash (Lineage)
        file_bytes = file_path.read_bytes()
        doc_hash = hashlib.sha256(file_bytes).hexdigest()
        
        # 2. Route based on type
        ext = file_path.suffix.lower()
        
        if ext in [".xlsx", ".xls", ".xlsm"]:
            return await self._parse_excel_defensive(file_path, doc_hash)
        else:
            # Word/PPT are safer, use standard MarkItDown
            return await self._parse_standard(file_path, doc_hash)

    async def _parse_excel_defensive(self, file_path: Path, doc_hash: str) -> List[IngestedChunk]:
        """
        The Defensive Strategy:
        1. Try MarkItDown (Best for structure/markdown).
        2. If fails/empty, fallback to OpenPyXL (Raw data).
        3. ALWAYS run Phantom Renderer for 'Shadow PDF' generation.
        """
        chunks: List[IngestedChunk] = []
        
        # Step A: The Phantom Pass (Visual Grounding)
        # We render the Excel to a PDF so the frontend has something to show.
        shadow_pdf_path = await self.phantom.render_excel_to_pdf(file_path)
        
        try:
            # Step B: Primary Parser (MarkItDown)
            logger.info("Attempting primary parse with MarkItDown...")
            result = self.md_converter.convert(str(file_path))
            
            if not result.text_content.strip():
                raise ValueError("MarkItDown returned empty content")
                
            # Create the chunk from Markdown
            chunks.append(self._create_chunk(
                text=result.text_content,
                doc_hash=doc_hash,
                source=file_path.name,
                shadow_path=shadow_pdf_path,
                method="markitdown"
            ))
            
        except Exception as e:
            # Step C: Fallback (OpenPyXL)
            logger.warning(f"⚠️ MarkItDown failed: {e}. Engaging OpenPyXL fallback.")
            raw_text = self._fallback_openpyxl(file_path)
            
            chunks.append(self._create_chunk(
                text=raw_text,
                doc_hash=doc_hash,
                source=file_path.name,
                shadow_path=shadow_pdf_path,
                method="openpyxl_fallback"
            ))
            
        return chunks

    async def _parse_standard(self, file_path: Path, doc_hash: str) -> List[IngestedChunk]:
        """Standard parsing for Word/PPT."""
        result = self.md_converter.convert(str(file_path))
        return [self._create_chunk(result.text_content, doc_hash, file_path.name)]

    def _fallback_openpyxl(self, file_path: Path) -> str:
        """
        Brute-force extraction of cell values. 
        Used when MarkItDown chokes on macros or binary formats.
        """
        wb = load_workbook(file_path, data_only=True)
        text_parts = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text_parts.append(f"### Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                # Join non-None cells
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                if row_text:
                    text_parts.append(row_text)
                    
        return "\n".join(text_parts)

    def _create_chunk(self, text: str, doc_hash: str, source: str, shadow_path: str = None, method: str = "standard") -> IngestedChunk:
        """Helper to build the Pydantic object."""
        chunk_id = str(uuid4())
        return IngestedChunk(
            chunk_id=chunk_id,
            doc_hash=doc_hash,
            clean_text=text.strip(), # We'll improve cleaning later
            raw_text=text,
            page_number=1, # Native files are treated as Stream 1
            token_count=len(text.split()),
            metadata={
                "source": source,
                "method": method,
                "shadow_pdf": shadow_path,
                "type": "native"
            }
        )